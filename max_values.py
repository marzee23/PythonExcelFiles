import os
import openpyxl
import pandas as pd

#Prompt for directory name to find max values in text files
directory = input("Please enter the file directory: ")

#Create empty excel workbook with 1 empty worksheet
#wb = openpyxl.Workbook('MaxValues.xlsx')
#ws = wb.create_sheet('Max Numbers')
#ws['A1'] = "File Name"
#ws['B1'] = 'Maximum Value'

def clean_line(line):
    return int(line.strip(';\n'))

#checks for existence of directory given by user
if os.path.lexists(directory):
    print(os.path.split(directory))
    max_numbers_dict = {}
    #iterates over files in given directory
    for path, dirs, files in os.walk(directory):
        #opens files one by one
        for blah in files:

            filepath = os.path.join(directory, blah)
            f = open(filepath)

            numbers = list(map(clean_line, f.readlines()))


            if len(numbers) > 0:
                print(max(numbers))
                max_numbers_dict.update({blah :max(numbers)})

            #for line in f.readlines():
            #    strip_line = line.strip(';\n')
            #    numbers.append(int(strip_line))
            else:

                print(filepath + "is empty.")

        #wb.save("MaxValues.xlsx")
        #wb.close("MaxValues.xlsx")
else:
    print("That pathway,"+ directory + " is no good.")

max_data_frame = pd.Series(max_numbers_dict)
print(max_data_frame)
